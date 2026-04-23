# contacts_handler.py - Обработчик контактов и справочника

import sqlite3
from contextlib import contextmanager

DATABASE = 'emails.db'

@contextmanager
def get_db():
    """Контекстный менеджер для БД"""
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    try:
        yield conn
    finally:
        conn.close()

def create_contacts_table():
    """Создать таблицу контактов если не существует"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS contacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT UNIQUE,
                full_name TEXT,
                position TEXT,
                department TEXT,
                phone TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        """)
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_email ON contacts(email)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_position ON contacts(position)')
        cursor.execute('CREATE INDEX IF NOT EXISTS idx_department ON contacts(department)')
        conn.commit()
        print("✅ Таблица contacts готова")

def add_contact(email, full_name='', position='', department='', phone=''):
    """Добавить или обновить контакт"""
    if not email:
        return False
    
    with get_db() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("""
                INSERT OR REPLACE INTO contacts 
                (email, full_name, position, department, phone)
                VALUES (?, ?, ?, ?, ?)
            """, (email.lower(), full_name, position, department, phone))
            conn.commit()
            return True
        except Exception as e:
            print(f"❌ Ошибка добавления контакта: {e}")
            return False

def get_contact(email):
    """Получить контакт по email"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, email, full_name, position, department, phone 
            FROM contacts 
            WHERE email = ?
        """, (email.lower(),))
        row = cursor.fetchone()
        
        if row:
            return dict(row)
        return None

def get_all_contacts():
    """Получить все контакты"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, email, full_name, position, department, phone 
            FROM contacts 
            ORDER BY full_name
        """)
        return [dict(row) for row in cursor.fetchall()]

def get_all_positions():
    """Получить все должности"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT position 
            FROM contacts 
            WHERE position IS NOT NULL AND position != ''
            ORDER BY position
        """)
        return [row['position'] for row in cursor.fetchall()]

def get_all_departments():
    """Получить все отделы"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT DISTINCT department 
            FROM contacts 
            WHERE department IS NOT NULL AND department != ''
            ORDER BY department
        """)
        return [row['department'] for row in cursor.fetchall()]

def load_from_xlsx(filepath):
    """Загрузить контакты из XLSX"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        added = 0
        for row in range(2, ws.max_row + 1):
            email = ws.cell(row, 4).value  # Column D - Email
            full_name = ws.cell(row, 1).value  # Column A - Full name
            position = ws.cell(row, 10).value  # Column J - Position
            department = ws.cell(row, 9).value  # Column I - Department
            phone = ws.cell(row, 6).value  # Column F - Phone
            
            if email:
                if add_contact(email, full_name or '', position or '', department or '', phone or ''):
                    added += 1
        
        print(f"✅ Загружено {added} контактов из XLSX")
        return added
    except Exception as e:
        print(f"❌ Ошибка загрузки XLSX: {e}")
        return 0

def delete_contact(email):
    """Удалить контакт"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute("DELETE FROM contacts WHERE email = ?", (email.lower(),))
        conn.commit()
        return cursor.rowcount > 0

# Инициализация при импорте
create_contacts_table()
